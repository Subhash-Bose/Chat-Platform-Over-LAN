# import socket library
import math
import random
import socket
from email import encoders
from email.mime.base import MIMEBase
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
import smtplib
import ssl
import time
import tqdm
# import threading library
import threading
import openpyxl as xl
# Choose a port that is free

import socket
import tqdm
SERVER = socket.gethostbyname(socket.gethostname())
SERVER="172.16.185.40"
PORT = 5000
file1 = open("backend//myfile.txt", "w")
L = ["chat histroy"]
file1.writelines(L)
file1.close()
while True:
	try:






		recivingTheattecedfile=" "
		def reciver(name,conn):
			try:
				print("recived file : ",name)
				while True:
					
					try:

						# client,addr=server.accept()
						file_name=conn.recv(1024).decode(FORMAT)
						print("Filename is",file_name)
						print("waiting for file size")
						recivingTheattecedfile=file_name
						file_size=conn.recv(1024).decode(FORMAT)
						print("and filesize",file_size)
					except:
						print("filename not recieved")
					
					# return

					file=open(file_name,"wb")
					
					file_bytes=b""
					done=False
					print("File opened to write")

					progress=tqdm.tqdm(unit="B",unit_scale=True,unit_divisor=1000,total=int(file_size))

					while not done:
						data=conn.recv(1024)
						print(" bits recieving")
						if file_bytes[-5:]==b"<END>":
							done=True
						else:
							file_bytes+=data
							progress.update(1024)

					file.write(file_bytes)
					file.close()
					print("file downloaded")
					break
			except:
				print("file not downloaded")
		# An IPv4 address is obtained
		# for the server.
		SERVER = socket.gethostbyname(socket.gethostname())
		# SERVER="172.16.185.40"
		# Address is stored as a tuple
		ADDRESS = (SERVER, PORT)

		# the format in which encoding
		# and decoding will occur
		FORMAT = "utf-8"

		clients, names = [], []
		server = socket.socket(socket.AF_INET,
							socket.SOCK_STREAM)

		# bind the address of the
		# server to the socket
		server.bind(ADDRESS)

		# function to start the connection


		def startChat():

			print("Success: Server Node address is " + SERVER)

			# listening for connections
			server.listen()

			while True:

				# accept connections and returns
				# a new connection to the client
				# and the address bound to it
				conn, addr = server.accept()
				# conn.send("NAME".encode(FORMAT))
				clients.append(conn)

				# while True:
				# 	msg conn.recv(1024).decode(FORMAT)
				# login_thread = threading.Thread(target=login,args=(conn, addr))
				print("Success: Client connected from address ",addr[0])
				# login_thread.start()
				try:
					handle_thread = threading.Thread(target=login,args=(conn, addr))
					handle_thread.start()
				except:
					pass
					
				# 1024 represents the max amount
				# of data that can be received (bytes)
				


		# 		# append the name and client
		# 		# to the respective list
		# 		names.append(name)
		# 		clients.append(conn)

		# 		print(f"Name is :{name}")

		# 		# broadcast message
		# 		broadcastMessage(f"{name} has joined the chat!".encode(FORMAT))

		# 		conn.send('Connection successful!'.encode(FORMAT))

		# 		# Start the handling thread
		# 		thread = threading.Thread(target=handle,
		# 								args=(conn, addr))
		# 		thread.start()

		# 		# no. of clients connected
		# 		# to the server
		# 		print(f"active connections {threading.activeCount()-1}")

		# # method to handle the
		# # incoming messages

		def login(conn,addr):
			global otp
			while True:
				try:
					msg=""
					try:
						msg=conn.recv(1024).decode(FORMAT)
					except:
						print("Error : Login Error")
						return
					if msg=="signup":
						try:
							msg = conn.recv(1024).decode(FORMAT)
						except:
							print("Error : Signup data not received")
						# print("signup request recieved for ",msg)

						signup_data=msg.split(":")
						# print(signup_data)
						try:
							try:
								wb=xl.load_workbook("backend\\login_data.xlsx")
							except:
								print("Error :File Error, backend login_data.xlsx missing, trying to create")
								wb=xl.Workbook()
								wb.save("backend\\login_data.xlsx")
								wb=xl.load_workbook("backend\\login_data.xlsx")
								print("Success: Backend login_data.xlsx created")
								
							# signup_data=["sjhbd","asd@jk.lk","ahjdkhj"]
							ws=wb.active
							nth_row=ws.max_row+1
							ws.cell(row=nth_row,column=1).value=signup_data[0].title()
							# print(signup_data[0]," has been written")
							ws.cell(row=nth_row,column=2).value=signup_data[1].lower()+"#"+signup_data[2]
							# print(signup_data[1].lower()+"#"+signup_data[2]," has been written")
							wb.save("backend\\login_data.xlsx")
							print("Success: Signup data has been stored")
						except:
							print("Error : Signup data not created")
						conn.send(str("Signup created for "+signup_data[0].title()).encode(FORMAT))

					if msg=="login":
						msg = conn.recv(1024).decode(FORMAT)
						login_data=msg.split(":")

						userid=login_data[0]+"#"+login_data[1]
						try:
							wb=xl.load_workbook("backend\\login_data.xlsx")
						except:
							msg=conn.send("failed".encode(FORMAT))
							return

						ws=wb.active
						found=0
						for nm,id in zip(ws["A"],ws["B"]):
							if id.value==userid:
								print("Success: ",nm.value,"Connected with IP address",addr[0])
								msg=conn.send(str("success"+str(nm.value)).encode(FORMAT))
								names.append(nm.value)
								active_names="#".join(names)
								chat_msg="client#"+nm.value
								for client in clients:
									try:
										client.send(chat_msg.encode(FORMAT))
										time.sleep(0.1)
										client.send(str("active#"+active_names).encode(FORMAT))
									except:
										pass

								handle(conn,addr,nm.value)
								return
						else:
							msg=conn.send("failed".encode(FORMAT))
					
					if msg=="reset":
						msg = conn.recv(1024).decode(FORMAT)
						email=msg
						print("Request: Reset request recieved from", email)
						otp=send_email(email)
						print("Success: OTP generated successfully ",otp)
						
					if msg=="requestotp":
						msg=conn.recv(1024).decode(FORMAT)
						# print("Recieved mail and otp is",msg)
						msg=msg.split("#")
						recieved_otp=msg[0]
						email=msg[1]
						# print("Recieved OTP is",recieved_otp,"from",email)
						# print("types are ",type(recieved_otp),recieved_otp,type(otp),otp)
						if(str(recieved_otp)==str(otp)):
							try:
								wb=xl.load_workbook("backend\\login_data.xlsx")
							except:
								msg="Failed: Email "+email+" Not found"
								msg=conn.send(msg.encode(FORMAT))
								return
							# print("database loaded")
							ws=wb.active
							for nm,id in zip(ws["A"],ws["B"]):
								pattern=id.value
								# print("str Pattern",pattern)
								if not pattern:
									continue
								pattern=pattern.split("#")
								# print("list Pattern",pattern)
								email_1=pattern[0]
								password=pattern[1]
								if email_1==email:
									print("Success: Email {} found in database".format(email))
									msg="Name: "+str(nm.value).title()+"\nEmail id: "+email+"\nPassword: "+password
									msg=conn.send(msg.encode(FORMAT))
									break
							else:
								msg="Failed: Email "+email+" Not found"
								msg=conn.send(msg.encode(FORMAT))
						else:
							msg="invalidotp"
							msg=conn.send(msg.encode(FORMAT))

				except:
					print("Error : Authentication Error")


		chat_msg=""
		def handle(conn, addr):
			
			# send_thread= threading.Thread(target=sendMsg,args=(conn, addr))
			recv_thread= threading.Thread(target=recvMsg,args=(conn, addr))
			# send_thread.start()
			recv_thread.start()

		def recvMsg(conn,addr,nm):
			# print("here")
			global chat_msg
			# print("your name is",nm)
			while True:
				try:
					# print("Waiting for message...")
					chat_msg=conn.recv(1024).decode(FORMAT)
					# print("recieved msg length=",len(chat_msg),chat_msg)
					if(chat_msg[:chat_msg.find("#")]=="ATTACHMENT"):
						print("Request: Attachment Request recieved")
						# print("waiting for file name")
						try:
							recv_thread2= threading.Thread(target=reciver,args=("rohit",conn))
							recv_thread2.start()
							recv_thread2.join()
						# continue
							# reciver("rohit",conn)
						except:
							print("Error : Attachment recieved failed")
						time.sleep(10)
					if chat_msg[:chat_msg.find("#")]=="exit":
						# print("recieved exit request")

						# print(chat_msg)
						# print("all names are",names)
						try:
							# print(str(conn))
							client_index=names.index(nm)
							# print("index is",client_index)
							# print(clients)
							# print()
							print("Info  : "+names[client_index],"left the chat")
							del names[client_index]

							# print(clients)
							del clients[client_index]
							# print("length of name and cl is",len(names))
							# print(conn)
							# print(clients)
							conn.close()
							print("Info  : Disconnected from ",addr[0])
							broadcastMessage("left#"+nm)
							for lost in names:
								broadcastMessage("active#"+lost)
								time.sleep(0.5)
							# print("msg broadcasted")
							# return
						except:
							# print("Error in cl")
							return
						return
					# for client in clients:
					# 	client.send(chat_msg.encode(FORMAT))
					broadcastMessage(chat_msg)
					# print("Sent msg is:",chat_msg)
					# print("Recieved msg is:",chat_msg)
				except:
					pass
		def broadcastMessage(message):
				
				for client in clients:
					try:
						client.send(message.encode(FORMAT))
						file1 = open("backend//myfile.txt", "a")  # append mode
						file1.write(message+"\n")
						file1.close()
					except:
						# print("Connection not found")
						pass
				
		def send_email(email):
			otp=random.randint(100000,999999)   
			# email="coolbose7@gmail.com"                                                                
			try:
				# Creating email
				subject = "Reset Password Requested"                                         
				body = "Your OTP to reset password is "+str(otp)
				sender_email = "coolbose7@gmail.com" 
				receiver_email = email
				pswd = "bjxjaihjhxljnmcy"

				# Create a multipart message and set headers
				msg=MIMEMultipart()                                                       
				msg["From"] = sender_email
				msg["To"] = receiver_email
				msg["Subject"] = subject

				# Add body to email
				msg.attach(MIMEText(body, "plain"))
				# convert message to string
				text = msg.as_string()

				# Log in to server using secure context and send email
				context = ssl.create_default_context()
				with smtplib.SMTP_SSL("smtp.gmail.com", 465, context=context) as server:
					server.login(sender_email, pswd)
					server.sendmail(sender_email, receiver_email, text)
			except:
				print("Error : OTP not sent")
			print("Success: Email sent to ",email)
			return otp
 
		startChat()
		
	except:
		exit()
